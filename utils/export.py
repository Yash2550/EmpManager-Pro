import os
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def export_employees_excel(employees):
    """Export employee list to an Excel file. Returns BytesIO buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ['Emp Code', 'Name', 'Email', 'Phone', 'Department',
               'Designation', 'Date of Joining', 'Salary', 'Status']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    for row, emp in enumerate(employees, 2):
        dept_name = emp.department.name if emp.department else 'N/A'
        values = [emp.emp_code, emp.full_name, emp.email, emp.phone or '',
                  dept_name, emp.designation or '', 
                  emp.date_of_joining.strftime('%d-%m-%Y') if emp.date_of_joining else '',
                  emp.salary, emp.status]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
    
    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 12)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_attendance_excel(records):
    """Export attendance records to Excel. Returns BytesIO buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ['Emp Code', 'Name', 'Date', 'Check In', 'Check Out', 'Hours', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    for row, rec in enumerate(records, 2):
        values = [
            rec.employee.emp_code, rec.employee.full_name,
            rec.date.strftime('%d-%m-%Y') if rec.date else '',
            rec.check_in.strftime('%I:%M %p') if rec.check_in else '',
            rec.check_out.strftime('%I:%M %p') if rec.check_out else '',
            f"{rec.hours_worked:.1f}" if rec.hours_worked else '0.0',
            rec.status
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
    
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 12)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_payroll_excel(payrolls):
    """Export payroll records to Excel. Returns BytesIO buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ['Emp Code', 'Name', 'Month', 'Year', 'Basic', 'HRA', 'DA',
               'Gross', 'PF', 'Tax', 'Net Salary', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    import calendar
    for row, p in enumerate(payrolls, 2):
        values = [
            p.employee.emp_code, p.employee.full_name,
            calendar.month_name[p.month], p.year,
            p.basic_salary, p.hra, p.da, p.gross_salary,
            p.pf_deduction, p.tax_deduction, p.net_salary, p.status
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
    
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 12)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
